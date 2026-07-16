from __future__ import annotations

import calendar
import json
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

MAX_INTERNAL_BREAK = timedelta(hours=1)
MAX_CONSECUTIVE_DAYS = 5
MAX_SHIFT_HOURS = 7.5
MIN_SHIFT_HOURS = 4.0
MIN_REST_HOURS = 11.0
WEEKLY_HOURS_TOLERANCE = 0.01

SCHEDULE_SOURCES: dict[str, str] = {
    "planned": "Plan publicado",
    "plannedDraft": "Borrador del planificador",
    "plannedDraftManuallyEdited": "Borrador editado manualmente",
}


@dataclass(frozen=True)
class ShiftRow:
    store_id: Any
    person_id: Any
    applicable_working_hours: Any
    work_day: date
    shift_start: datetime
    shift_end: datetime
    worked_hours: float
    break_hours: float


@dataclass(frozen=True)
class AbsenceDay:
    store_id: Any
    person_id: Any
    absence_day: date
    absence_type: str
    absence_status: str


@dataclass(frozen=True)
class Incident:
    store_id: Any
    person_id: Any
    month: str
    incident_type: str
    start_date: date
    end_date: date
    observed_value: float
    limit_text: str
    detail: str


@dataclass
class ValidationResult:
    source_data: dict[str, Any]
    schedule_source: str
    shifts: list[ShiftRow]
    employee_months: dict[tuple[Any, Any, str], Any]
    absences: list[AbsenceDay]
    employee_presence_dates: dict[tuple[Any, Any], set[date]]
    summaries: list[dict[str, Any]]
    incidents: list[Incident]
    weekly_rows: list[dict[str, Any]]
    data_dates: set[date]


def validate_schedule_source(schedule_source: str) -> str:
    if schedule_source not in SCHEDULE_SOURCES:
        valid = ", ".join(SCHEDULE_SOURCES)
        raise ValueError(f"Origen de horarios no valido: {schedule_source!r}. Valores: {valid}")
    return schedule_source


def parse_iso_datetime(value: str) -> datetime:
    if not isinstance(value, str) or not value.strip():
        raise ValueError(f"Fecha no valida: {value!r}")
    normalized = value.strip()
    if normalized.endswith("Z"):
        normalized = normalized[:-1] + "+00:00"
    return datetime.fromisoformat(normalized).replace(tzinfo=None)


def load_json_bytes(file_bytes: bytes) -> dict[str, Any]:
    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("El fichero no esta codificado en UTF-8.") from exc
    try:
        data = json.loads(text)
    except json.JSONDecodeError as exc:
        raise ValueError(
            f"JSON no valido. Linea {exc.lineno}, columna {exc.colno}: {exc.msg}"
        ) from exc
    if not isinstance(data, dict):
        raise ValueError("La raiz del JSON debe ser un objeto.")
    return data


def load_json_path(path: Path) -> dict[str, Any]:
    return load_json_bytes(path.read_bytes())


def detect_schedule_sources(data: dict[str, Any]) -> dict[str, dict[str, int]]:
    """Cuenta dias-persona y segmentos disponibles para cada origen de horarios."""
    result = {
        key: {"person_days": 0, "segments": 0, "work_segments": 0}
        for key in SCHEDULE_SOURCES
    }
    for store_day in data.get("storeDayTimes") or []:
        if not isinstance(store_day, dict):
            continue
        for person_day in store_day.get("people") or []:
            if not isinstance(person_day, dict):
                continue
            day_times = person_day.get("dayTimes") or {}
            for source in SCHEDULE_SOURCES:
                segments = day_times.get(source)
                if not isinstance(segments, list) or not segments:
                    continue
                result[source]["person_days"] += 1
                result[source]["segments"] += len(segments)
                result[source]["work_segments"] += sum(
                    1
                    for segment in segments
                    if isinstance(segment, dict)
                    and str(segment.get("hourType", "")).upper() == "WORK"
                )
    return result


def month_key(day: date) -> str:
    return day.strftime("%Y-%m")


def extract_data(data: dict[str, Any], schedule_source: str):
    schedule_source = validate_schedule_source(schedule_source)
    store_id = (data.get("store") or {}).get("id")
    store_day_times = data.get("storeDayTimes") or []
    if not isinstance(store_day_times, list):
        raise ValueError("El campo 'storeDayTimes' debe ser una lista.")

    shifts: list[ShiftRow] = []
    employee_months: dict[tuple[Any, Any, str], Any] = {}
    absences: list[AbsenceDay] = []
    employee_presence_dates: dict[tuple[Any, Any], set[date]] = defaultdict(set)

    for store_day in store_day_times:
        if not isinstance(store_day, dict) or not store_day.get("operatingDate"):
            continue
        operating_day = date.fromisoformat(str(store_day["operatingDate"])[:10])
        for person_day in store_day.get("people") or []:
            if not isinstance(person_day, dict):
                continue
            person = person_day.get("person") or {}
            person_id = person_day.get("personId", person.get("personId"))
            applicable_hours = person.get("applicableWorkingHours")
            employee_presence_dates[(store_id, person_id)].add(operating_day)
            employee_months[(store_id, person_id, month_key(operating_day))] = applicable_hours

            day_times = person_day.get("dayTimes") or {}
            selected_schedule = day_times.get(schedule_source) or []
            if not isinstance(selected_schedule, list):
                selected_schedule = []

            seen_absences: set[tuple[str, str]] = set()
            for absence in day_times.get("absences") or []:
                if not isinstance(absence, dict):
                    continue
                status = str(absence.get("status") or "").upper()
                if status not in {"VALIDATED", "APPROVED"}:
                    continue
                type_data = absence.get("type") or {}
                absence_type = str(
                    type_data.get("name")
                    or type_data.get("description")
                    or absence.get("id")
                    or "AUSENCIA"
                )
                key = (absence_type, status)
                if key in seen_absences:
                    continue
                seen_absences.add(key)
                absences.append(
                    AbsenceDay(store_id, person_id, operating_day, absence_type, status)
                )

            segments: list[tuple[datetime, datetime]] = []
            for segment in selected_schedule:
                if not isinstance(segment, dict):
                    continue
                if str(segment.get("hourType", "")).upper() != "WORK":
                    continue
                start_value = segment.get("startDateTime")
                end_value = segment.get("endDateTime")
                if not start_value or not end_value:
                    continue
                start_dt = parse_iso_datetime(start_value)
                end_dt = parse_iso_datetime(end_value)
                if end_dt <= start_dt:
                    raise ValueError(
                        f"Segmento invalido en {schedule_source}: personId={person_id}, "
                        f"inicio={start_value}, fin={end_value}"
                    )
                segments.append((start_dt, end_dt))

            if not segments:
                continue
            segments.sort(key=lambda item: item[0])
            shift_start = segments[0][0]
            shift_end = max(end for _, end in segments)
            net_work = sum((end - start for start, end in segments), timedelta())
            break_duration = timedelta()
            previous_end = segments[0][1]
            for current_start, current_end in segments[1:]:
                gap = current_start - previous_end
                if timedelta(0) < gap <= MAX_INTERNAL_BREAK:
                    break_duration += gap
                if current_end > previous_end:
                    previous_end = current_end
            shifts.append(
                ShiftRow(
                    store_id,
                    person_id,
                    applicable_hours,
                    operating_day,
                    shift_start,
                    shift_end,
                    round(net_work.total_seconds() / 3600, 4),
                    round(break_duration.total_seconds() / 3600, 4),
                )
            )

    shifts.sort(key=lambda row: (str(row.store_id), str(row.person_id), row.work_day))
    absences.sort(key=lambda row: (str(row.store_id), str(row.person_id), row.absence_day))
    return shifts, employee_months, absences, employee_presence_dates


def daterange(start: date, end: date) -> Iterable[date]:
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


def find_consecutive_streaks(days: list[date]) -> list[list[date]]:
    if not days:
        return []
    unique_days = sorted(set(days))
    streaks = [[unique_days[0]]]
    for current in unique_days[1:]:
        if current == streaks[-1][-1] + timedelta(days=1):
            streaks[-1].append(current)
        else:
            streaks.append([current])
    return streaks


def weekend_counts(year: int, month: int, worked_days: set[date]):
    month_days = [date(year, month, d) for d in range(1, calendar.monthrange(year, month)[1] + 1)]
    saturdays = [d for d in month_days if d.weekday() == 5]
    sundays = [d for d in month_days if d.weekday() == 6]
    free_saturdays = sum(d not in worked_days for d in saturdays)
    free_sundays = sum(d not in worked_days for d in sundays)
    complete = sum(
        saturday not in worked_days
        and (saturday + timedelta(days=1)).month == month
        and saturday + timedelta(days=1) not in worked_days
        for saturday in saturdays
    )
    return complete, free_saturdays, free_sundays


def week_start(day: date) -> date:
    return day - timedelta(days=day.weekday())


def analyze_weekly_hours(shifts, employee_months, data_dates, absences, employee_presence_dates):
    hours_by_week = defaultdict(float)
    hours_by_employee = defaultdict(list)
    worked_days = defaultdict(set)
    applicable_by_employee = {}
    absences_by_day = defaultdict(set)

    for (store_id, person_id, _), applicable in employee_months.items():
        applicable_by_employee[(store_id, person_id)] = applicable
    for shift in shifts:
        employee = (shift.store_id, shift.person_id)
        hours_by_week[(shift.store_id, shift.person_id, week_start(shift.work_day))] += shift.worked_hours
        hours_by_employee[employee].append(shift.worked_hours)
        worked_days[employee].add(shift.work_day)
        if shift.applicable_working_hours not in (None, ""):
            applicable_by_employee[employee] = shift.applicable_working_hours
    for absence in absences:
        absences_by_day[(absence.store_id, absence.person_id, absence.absence_day)].add(absence.absence_type)
    if not data_dates:
        return []

    first_week = week_start(min(data_dates))
    last_week = week_start(max(data_dates))
    week_starts = [d for d in daterange(first_week, last_week) if d.weekday() == 0]
    rows = []
    for (store_id, person_id), applicable in sorted(
        applicable_by_employee.items(), key=lambda item: (str(item[0][0]), str(item[0][1]))
    ):
        employee = (store_id, person_id)
        values = hours_by_employee.get(employee, [])
        average_daily = round(sum(values) / len(values), 4) if values else None
        presence_dates = employee_presence_dates.get(employee, set())
        all_absence_dates = {
            day for (sid, pid, day), types in absences_by_day.items()
            if sid == store_id and pid == person_id and types
        }
        absent_entire_period = bool(presence_dates and presence_dates.issubset(all_absence_dates) and not worked_days.get(employee))

        for monday in week_starts:
            sunday = monday + timedelta(days=6)
            week_days = {monday + timedelta(days=i) for i in range(7)}
            covered = len(week_days & data_dates)
            complete = covered == 7
            planned = round(hours_by_week.get((store_id, person_id, monday), 0.0), 4)
            try:
                contracted = float(applicable)
            except (TypeError, ValueError):
                contracted = None
            difference = round(planned - contracted, 4) if contracted is not None else None
            missing = round(max(contracted - planned, 0.0), 4) if contracted is not None else None
            excess = round(max(planned - contracted, 0.0), 4) if contracted is not None else None
            absence_days = sorted(
                d for d in week_days
                if (store_id, person_id, d) in absences_by_day
                and d not in worked_days.get(employee, set())
                and d in data_dates
            )
            absence_types = sorted({t for d in absence_days for t in absences_by_day[(store_id, person_id, d)]})
            potential = round(len(absence_days) * average_daily, 4) if average_daily is not None else None
            if not complete:
                status = "NO EVALUABLE"
            elif contracted is None:
                status = "SIN HORAS CONTRATO"
            elif abs(difference) <= WEEKLY_HOURS_TOLERANCE:
                status = "COINCIDE"
            elif difference < 0:
                status = "FALTAN HORAS"
            else:
                status = "EXCESO HORAS"
            missing_and_absence = bool(complete and missing is not None and missing > WEEKLY_HOURS_TOLERANCE and absence_days)
            if absent_entire_period:
                explanation = "AUSENTE TODO EL PERIODO"
            elif not missing_and_absence:
                explanation = "NO"
            elif potential is None:
                explanation = "AUSENCIA SIN MEDIA CALCULABLE"
            elif potential + WEEKLY_HOURS_TOLERANCE >= missing:
                explanation = "PODRIA EXPLICAR TODAS LAS HORAS FALTANTES"
            else:
                explanation = "PODRIA EXPLICAR PARTE DE LAS HORAS FALTANTES"
            rows.append({
                "id_tienda": store_id,
                "personId": person_id,
                "ano_iso": monday.isocalendar().year,
                "semana_iso": monday.isocalendar().week,
                "inicio_semana": monday,
                "fin_semana": sunday,
                "dias_cubiertos_fichero": covered,
                "semana_completa_en_fichero": "SI" if complete else "NO",
                "applicableWorkingHours": contracted,
                "horas_planificadas": planned,
                "diferencia_planificadas_menos_contrato": difference,
                "horas_no_planificadas_hasta_contrato": missing,
                "horas_planificadas_en_exceso": excess,
                "estado_planificacion": status,
                "cumple_horas_contrato": "SI" if status == "COINCIDE" else "NO" if status in {"FALTAN HORAS", "EXCESO HORAS"} else "NO EVALUABLE",
                "media_horas_dia_planificado": average_daily,
                "dias_ausencia_sin_turno": len(absence_days),
                "fechas_ausencia_sin_turno": ", ".join(d.isoformat() for d in absence_days),
                "tipos_ausencia": ", ".join(absence_types),
                "horas_potenciales_asociadas_ausencia": potential,
                "faltan_horas_y_hay_ausencia": "SI" if missing_and_absence else "NO",
                "posible_explicacion_por_ausencia": explanation,
                "ausente_todo_el_periodo": "SI" if absent_entire_period else "NO",
            })
    return rows


def analyze_shifts(shifts, employee_months):
    by_employee = defaultdict(list)
    for shift in shifts:
        by_employee[(shift.store_id, shift.person_id)].append(shift)
    incidents: list[Incident] = []
    for (store_id, person_id), person_shifts in by_employee.items():
        person_shifts.sort(key=lambda row: (row.shift_start, row.shift_end))
        for shift in person_shifts:
            month = month_key(shift.work_day)
            if shift.worked_hours > MAX_SHIFT_HOURS:
                incidents.append(Incident(store_id, person_id, month, "TURNO_SUPERIOR_7_5H", shift.work_day, shift.work_day, shift.worked_hours, "<= 7,5 horas", f"{shift.work_day:%d/%m/%Y}: {shift.worked_hours:.2f} h ({shift.shift_start:%H:%M}-{shift.shift_end:%H:%M})"))
            if shift.worked_hours < MIN_SHIFT_HOURS:
                incidents.append(Incident(store_id, person_id, month, "TURNO_INFERIOR_4H", shift.work_day, shift.work_day, shift.worked_hours, ">= 4 horas", f"{shift.work_day:%d/%m/%Y}: {shift.worked_hours:.2f} h ({shift.shift_start:%H:%M}-{shift.shift_end:%H:%M})"))
        for previous, current in zip(person_shifts, person_shifts[1:]):
            rest = (current.shift_start - previous.shift_end).total_seconds() / 3600
            if rest < MIN_REST_HOURS:
                incidents.append(Incident(store_id, person_id, month_key(current.work_day), "DESCANSO_INFERIOR_11H", previous.work_day, current.work_day, round(rest, 4), ">= 11 horas", f"{previous.work_day:%d/%m/%Y} {previous.shift_end:%H:%M} -> {current.work_day:%d/%m/%Y} {current.shift_start:%H:%M}: {rest:.2f} h"))
        for streak in find_consecutive_streaks([row.work_day for row in person_shifts]):
            if len(streak) > MAX_CONSECUTIVE_DAYS:
                for month in sorted({month_key(day) for day in streak}):
                    incidents.append(Incident(store_id, person_id, month, "MAS_DE_5_DIAS_CONSECUTIVOS", streak[0], streak[-1], float(len(streak)), "<= 5 dias", f"{streak[0]:%d/%m/%Y}-{streak[-1]:%d/%m/%Y}: {len(streak)} dias consecutivos"))

    groups = defaultdict(list)
    for incident in incidents:
        groups[(incident.store_id, incident.person_id, incident.month, incident.incident_type)].append(incident)
    summaries = []
    for (store_id, person_id, month), applicable in sorted(employee_months.items(), key=lambda item: (str(item[0][0]), str(item[0][1]), item[0][2])):
        year, month_number = map(int, month.split("-"))
        month_shifts = [row for row in by_employee.get((store_id, person_id), []) if month_key(row.work_day) == month]
        worked = {row.work_day for row in month_shifts}
        complete_weekends, free_saturdays, free_sundays = weekend_counts(year, month_number, worked)
        touching = [streak for streak in find_consecutive_streaks([row.work_day for row in by_employee.get((store_id, person_id), [])]) if any(month_key(day) == month for day in streak)]
        max_streak = max((len(streak) for streak in touching), default=0)
        count_consecutive = len(groups.get((store_id, person_id, month, "MAS_DE_5_DIAS_CONSECUTIVOS"), []))
        count_long = len(groups.get((store_id, person_id, month, "TURNO_SUPERIOR_7_5H"), []))
        count_short = len(groups.get((store_id, person_id, month, "TURNO_INFERIOR_4H"), []))
        count_rest = len(groups.get((store_id, person_id, month, "DESCANSO_INFERIOR_11H"), []))
        summaries.append({
            "id_tienda": store_id, "personId": person_id, "applicableWorkingHours": applicable, "mes": month,
            "dias_trabajados": len(worked), "max_dias_consecutivos": max_streak,
            "incidencias_dias_consecutivos": count_consecutive, "cumple_max_5_dias": "SI" if count_consecutive == 0 else "NO",
            "turnos_superiores_7_5h": count_long, "cumple_duracion_maxima": "SI" if count_long == 0 else "NO",
            "turnos_inferiores_4h": count_short, "cumple_duracion_minima": "SI" if count_short == 0 else "NO",
            "descansos_inferiores_11h": count_rest, "cumple_descanso_entre_jornadas": "SI" if count_rest == 0 else "NO",
            "cumple_todas_las_reglas": "SI" if count_consecutive + count_long + count_short + count_rest == 0 else "NO",
            "fines_semana_completos_libres": complete_weekends, "sabados_libres": free_saturdays, "domingos_libres": free_sundays,
        })
    incidents.sort(key=lambda item: (str(item.store_id), str(item.person_id), item.month, item.start_date, item.incident_type))
    return summaries, incidents


def collect_data_dates(data):
    return {
        date.fromisoformat(str(item.get("operatingDate"))[:10])
        for item in data.get("storeDayTimes") or []
        if isinstance(item, dict) and item.get("operatingDate")
    }


def run_validation(data: dict[str, Any], schedule_source: str = "plannedDraft") -> ValidationResult:
    validate_schedule_source(schedule_source)
    shifts, employee_months, absences, presence = extract_data(data, schedule_source)
    summaries, incidents = analyze_shifts(shifts, employee_months)
    dates = collect_data_dates(data)
    weekly = analyze_weekly_hours(shifts, employee_months, dates, absences, presence)
    return ValidationResult(data, schedule_source, shifts, employee_months, absences, presence, summaries, incidents, weekly, dates)


def result_dataframes(result: ValidationResult):
    shifts = pd.DataFrame([{
        "id_tienda": s.store_id, "personId": s.person_id, "applicableWorkingHours": s.applicable_working_hours,
        "day": s.work_day, "hora_inicio": s.shift_start, "hora_fin": s.shift_end,
        "horas_totales": s.worked_hours, "duracion_descanso": s.break_hours,
    } for s in result.shifts])
    incidents = pd.DataFrame([{
        "id_tienda": i.store_id, "personId": i.person_id, "mes": i.month, "tipo_incidencia": i.incident_type,
        "fecha_inicio": i.start_date, "fecha_fin": i.end_date, "valor_observado": i.observed_value,
        "limite": i.limit_text, "detalle": i.detail,
    } for i in result.incidents])
    absences = pd.DataFrame([{
        "id_tienda": a.store_id, "personId": a.person_id, "fecha": a.absence_day,
        "tipo_ausencia": a.absence_type, "estado": a.absence_status,
    } for a in result.absences])
    return {
        "shifts": shifts,
        "summaries": pd.DataFrame(result.summaries),
        "incidents": incidents,
        "weekly": pd.DataFrame(result.weekly_rows),
        "absences": absences,
    }


def build_excel_bytes(result: ValidationResult) -> bytes:
    frames = result_dataframes(result)
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheet_map = {
        "shifts": "Turnos",
        "summaries": "Validacion mensual",
        "incidents": "Detalle incidencias",
        "weekly": "Control horas semanal",
        "absences": "Ausencias",
    }
    for key, sheet_name in sheet_map.items():
        ws = workbook.create_sheet(sheet_name)
        frame = frames[key]
        if frame.empty:
            ws.append(["Sin registros"])
            continue
        ws.append(list(frame.columns))
        for row in frame.itertuples(index=False, name=None):
            ws.append(list(row))
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        table = Table(displayName=f"Table_{key}", ref=ws.dimensions)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
        for column_cells in ws.columns:
            width = min(max(len(str(c.value or "")) for c in column_cells) + 2, 45)
            ws.column_dimensions[column_cells[0].column_letter].width = width
    info = workbook.create_sheet("Informacion", 0)
    info.append(["Origen analizado", result.schedule_source])
    info.append(["Descripcion", SCHEDULE_SOURCES[result.schedule_source]])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
