from __future__ import annotations

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from workforce_validator.dataframes import result_dataframes
from workforce_validator.models import ValidationResult
from workforce_validator.schedule_sources import SCHEDULE_SOURCES


def build_excel_bytes(result: ValidationResult) -> bytes:
    frames = result_dataframes(result)
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheet_map = {
        "shifts": "Turnos",
        "summaries": "Validacion mensual",
        "incidents": "Detalle incidencias",
        "weekly": "Control horas semanal",
        "absences": "Ausencias",
    }
    for key, sheet_name in sheet_map.items():
        worksheet = workbook.create_sheet(sheet_name)
        frame = frames[key]
        if frame.empty:
            worksheet.append(["Sin registros"])
            continue
        worksheet.append(list(frame.columns))
        for row in frame.itertuples(index=False, name=None):
            worksheet.append(list(row))
        for cell in worksheet[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        table = Table(displayName=f"Table_{key}", ref=worksheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        worksheet.add_table(table)
        for column_cells in worksheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 45)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    info = workbook.create_sheet("Informacion", 0)
    info.append(["Origen analizado", result.schedule_source])
    info.append(["Descripcion", SCHEDULE_SOURCES[result.schedule_source]])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
