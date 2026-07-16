from collections import Counter
from datetime import date, timedelta
from openpyxl import load_workbook
from schedule_adapter import build_excel_bytes, run_validation


def seg(day,start,end):
    return {"hourType":"WORK","startDateTime":f"{day}T{start}:00","endDateTime":f"{day}T{end}:00"}

def regression_data():
    days=[]
    start=date(2026,8,3)
    for i in range(7):
        day=(start+timedelta(days=i)).isoformat()
        draft=[]
        published=[]
        absences=[]
        if i < 6:
            published=[seg(day,"09:00","17:00")]
            draft=[seg(day,"09:00","17:00")] if i == 0 else [seg(day,"09:00","15:00")]
        if i == 6:
            absences=[{"status":"VALIDATED","type":{"name":"VACACIONES"}}]
        days.append({"operatingDate":day,"people":[{"personId":"E1","person":{"personId":"E1","applicableWorkingHours":40},"dayTimes":{"planned":published,"plannedDraft":draft,"plannedDraftManuallyEdited":True,"absences":absences}}]})
    return {"store":{"id":14947},"storeDayTimes":days}

def test_end_to_end_snapshot(tmp_path):
    result=run_validation(regression_data(),"plannedDraft","all")
    assert len(result.shifts) == 6
    assert len(result.absences) == 1
    assert Counter(i.incident_type for i in result.incidents) == {"TURNO_SUPERIOR_7_5H":1,"MAS_DE_5_DIAS_CONSECUTIVOS":1}
    assert result.summaries[0]["max_dias_consecutivos"] == 6
    assert result.summaries[0]["cumple_todas_las_reglas"] == "NO"
    assert result.weekly_rows[0]["estado_planificacion"] == "FALTAN HORAS"
    assert result.weekly_rows[0]["horas_planificadas"] == 38.0
    output=tmp_path/"validation.xlsx"; output.write_bytes(build_excel_bytes(result))
    workbook=load_workbook(output,read_only=True)
    assert workbook.sheetnames == ["Informacion","Turnos","Validacion mensual","Detalle incidencias","Control horas semanal","Ausencias"]
