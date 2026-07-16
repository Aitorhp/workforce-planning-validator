from __future__ import annotations

import argparse
import calendar
import json
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


DEFAULT_INPUT = "JSON_RESPONSE_AUGUST_14947.txt"
DEFAULT_OUTPUT = "planned_drafts_validado.xlsx"

MAX_INTERNAL_BREAK = timedelta(hours=1)
MAX_CONSECUTIVE_DAYS = 5
MAX_SHIFT_HOURS = 7.5
MIN_SHIFT_HOURS = 4.0
MIN_REST_HOURS = 11.0
WEEKLY_HOURS_TOLERANCE = 0.01


@dataclass(frozen=True)
class ShiftRow:
    store_id: Any
    person_id: Any
    applicable_working_hours: Any
    work_day: date
    shift_start: datetime
    shift_end: datetime
    worked_hours: float
    break_hours: float


@dataclass(frozen=True)
class AbsenceDay:
    store_id: Any
    person_id: Any
    absence_day: date
    absence_type: str
    absence_status: str


@dataclass(frozen=True)
class Incident:
    store_id: Any
    person_id: Any
    month: str
    incident_type: str
    start_date: date
    end_date: date
    observed_value: float
    limit_text: str
    detail: str


def parse_iso_datetime(value: str) -> datetime:
    """Convierte una fecha ISO 8601 del JSON en datetime."""
    if not isinstance(value, str) or not value.strip():
        raise ValueError(f"Fecha no valida: {value!r}")

    normalized = value.strip()
    if normalized.endswith("Z"):
        normalized = normalized[:-1] + "+00:00"

    parsed = datetime.fromisoformat(normalized)
    # Se elimina la zona horaria para que Excel y las comparaciones sean simples.
    return parsed.replace(tzinfo=None)


def parse_operating_date(value: Any, fallback: date) -> date:
    if value in (None, ""):
        return fallback
    return date.fromisoformat(str(value)[:10])


def load_json(input_path: Path) -> dict[str, Any]:
    """Carga el JSON aunque el fichero tenga extension .txt o BOM UTF-8."""
    try:
        with input_path.open("r", encoding="utf-8-sig") as file:
            data = json.load(file)
    except FileNotFoundError as exc:
        raise FileNotFoundError(
            f"No se encontro el fichero de entrada: {input_path}"
        ) from exc
    except json.JSONDecodeError as exc:
        raise ValueError(
            f"El fichero no contiene un JSON valido. Linea {exc.lineno}, "
            f"columna {exc.colno}: {exc.msg}"
        ) from exc

    if not isinstance(data, dict):
        raise ValueError("La raiz del JSON debe ser un objeto.")
    return data


def month_key(day: date) -> str:
    return day.strftime("%Y-%m")


def extract_data(
    data: dict[str, Any],
) -> tuple[
    list[ShiftRow],
    dict[tuple[Any, Any, str], Any],
    list[AbsenceDay],
    dict[tuple[Any, Any], set[date]],
]:
    """
    Devuelve:
    - Turnos aplanados: un registro por tienda, empleado y dia.
    - Empleados presentes por mes, incluso si no tienen turno.
    - Ausencias diarias validadas detectadas en cada fecha del fichero.
    - Fechas en las que cada empleado aparece dentro del horizonte.

    Solo se consideran segmentos plannedDraft con hourType == WORK.
    Las horas trabajadas son la suma neta de los segmentos.
    Los huecos positivos de hasta una hora se consideran descanso interno.
    """
    store = data.get("store") or {}
    store_id = store.get("id")
    store_day_times = data.get("storeDayTimes") or []

    if not isinstance(store_day_times, list):
        raise ValueError("El campo 'storeDayTimes' debe ser una lista.")

    shifts: list[ShiftRow] = []
    employee_months: dict[tuple[Any, Any, str], Any] = {}
    absences: list[AbsenceDay] = []
    employee_presence_dates: dict[tuple[Any, Any], set[date]] = defaultdict(set)

    for store_day in store_day_times:
        if not isinstance(store_day, dict):
            continue

        operating_date_raw = store_day.get("operatingDate")
        if not operating_date_raw:
            continue
        operating_day = date.fromisoformat(str(operating_date_raw)[:10])
        people = store_day.get("people") or []
        if not isinstance(people, list):
            continue

        for person_day in people:
            if not isinstance(person_day, dict):
                continue

            person = person_day.get("person") or {}
            person_id = person_day.get("personId", person.get("personId"))
            applicable_hours = person.get("applicableWorkingHours")
            employee_presence_dates[(store_id, person_id)].add(operating_day)

            day_times = person_day.get("dayTimes") or {}
            planned_draft = day_times.get("plannedDraft") or []
            if not isinstance(planned_draft, list):
                planned_draft = []

            day_absences = day_times.get("absences") or []
            if isinstance(day_absences, list):
                seen_absences: set[tuple[str, str]] = set()
                for absence in day_absences:
                    if not isinstance(absence, dict):
                        continue
                    status = str(absence.get("status") or "").upper()
                    if status not in {"VALIDATED", "APPROVED"}:
                        continue
                    absence_type_data = absence.get("type") or {}
                    absence_type = str(
                        absence_type_data.get("name")
                        or absence_type_data.get("description")
                        or absence.get("id")
                        or "AUSENCIA"
                    )
                    absence_key = (absence_type, status)
                    if absence_key in seen_absences:
                        continue
                    seen_absences.add(absence_key)
                    absences.append(
                        AbsenceDay(
                            store_id=store_id,
                            person_id=person_id,
                            absence_day=operating_day,
                            absence_type=absence_type,
                            absence_status=status,
                        )
                    )

            segments: list[tuple[datetime, datetime]] = []
            for segment in planned_draft:
                if not isinstance(segment, dict):
                    continue
                if str(segment.get("hourType", "")).upper() != "WORK":
                    continue

                start_value = segment.get("startDateTime")
                end_value = segment.get("endDateTime")
                if not start_value or not end_value:
                    continue

                start_dt = parse_iso_datetime(start_value)
                end_dt = parse_iso_datetime(end_value)
                if end_dt <= start_dt:
                    raise ValueError(
                        "Segmento con fin anterior o igual al inicio: "
                        f"personId={person_id}, inicio={start_value}, fin={end_value}"
                    )
                segments.append((start_dt, end_dt))

            work_day = parse_operating_date(operating_date_raw, operating_day)
            employee_months[(store_id, person_id, month_key(work_day))] = applicable_hours

            if not segments:
                continue

            segments.sort(key=lambda item: item[0])
            shift_start = segments[0][0]
            shift_end = max(end for _, end in segments)
            net_work = sum((end - start for start, end in segments), timedelta())

            break_duration = timedelta()
            previous_end = segments[0][1]
            for current_start, current_end in segments[1:]:
                gap = current_start - previous_end
                if timedelta(0) < gap <= MAX_INTERNAL_BREAK:
                    break_duration += gap
                if current_end > previous_end:
                    previous_end = current_end

            shifts.append(
                ShiftRow(
                    store_id=store_id,
                    person_id=person_id,
                    applicable_working_hours=applicable_hours,
                    work_day=work_day,
                    shift_start=shift_start,
                    shift_end=shift_end,
                    worked_hours=round(net_work.total_seconds() / 3600, 4),
                    break_hours=round(break_duration.total_seconds() / 3600, 4),
                )
            )

    shifts.sort(key=lambda row: (str(row.store_id), str(row.person_id), row.work_day))
    absences.sort(key=lambda row: (str(row.store_id), str(row.person_id), row.absence_day))
    return shifts, employee_months, absences, employee_presence_dates

def daterange(start: date, end: date) -> Iterable[date]:
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


def find_consecutive_streaks(days: list[date]) -> list[list[date]]:
    if not days:
        return []

    unique_days = sorted(set(days))
    streaks: list[list[date]] = [[unique_days[0]]]
    for current in unique_days[1:]:
        if current == streaks[-1][-1] + timedelta(days=1):
            streaks[-1].append(current)
        else:
            streaks.append([current])
    return streaks


def weekend_counts(year: int, month: int, worked_days: set[date]) -> tuple[int, int, int]:
    """
    Cuenta sabados y domingos libres del mes.
    Un fin de semana completo solo cuenta si sabado y domingo pertenecen al
    mismo mes y ambos estan libres.
    """
    last_day = calendar.monthrange(year, month)[1]
    month_days = [date(year, month, day) for day in range(1, last_day + 1)]

    saturdays = [day for day in month_days if day.weekday() == 5]
    sundays = [day for day in month_days if day.weekday() == 6]

    free_saturdays = sum(day not in worked_days for day in saturdays)
    free_sundays = sum(day not in worked_days for day in sundays)

    complete_free_weekends = 0
    for saturday in saturdays:
        sunday = saturday + timedelta(days=1)
        if sunday.month != month:
            continue
        if saturday not in worked_days and sunday not in worked_days:
            complete_free_weekends += 1

    return complete_free_weekends, free_saturdays, free_sundays



def week_start(day: date) -> date:
    """Devuelve el lunes de la semana ISO de la fecha."""
    return day - timedelta(days=day.weekday())


def analyze_weekly_hours(
    shifts: list[ShiftRow],
    employee_months: dict[tuple[Any, Any, str], Any],
    data_dates: set[date],
    absences: list[AbsenceDay],
    employee_presence_dates: dict[tuple[Any, Any], set[date]],
) -> list[dict[str, Any]]:
    """
    Compara las horas planificadas con las horas contractuales semanales.

    Las ausencias NO se suman a las horas planificadas. Se muestran de forma
    independiente para identificar cuantas horas no planificadas podrian estar
    potencialmente explicadas por dias con ausencia.
    """
    hours_by_employee_week: dict[tuple[Any, Any, date], float] = defaultdict(float)
    worked_hours_by_employee: dict[tuple[Any, Any], list[float]] = defaultdict(list)
    worked_days_by_employee: dict[tuple[Any, Any], set[date]] = defaultdict(set)
    applicable_by_employee: dict[tuple[Any, Any], Any] = {}
    absence_types_by_employee_day: dict[tuple[Any, Any, date], set[str]] = defaultdict(set)

    for (store_id, person_id, _month), applicable in employee_months.items():
        applicable_by_employee[(store_id, person_id)] = applicable

    for shift in shifts:
        employee_key = (shift.store_id, shift.person_id)
        key = (shift.store_id, shift.person_id, week_start(shift.work_day))
        hours_by_employee_week[key] += shift.worked_hours
        worked_hours_by_employee[employee_key].append(shift.worked_hours)
        worked_days_by_employee[employee_key].add(shift.work_day)
        if shift.applicable_working_hours not in (None, ""):
            applicable_by_employee[employee_key] = shift.applicable_working_hours

    for absence in absences:
        absence_types_by_employee_day[
            (absence.store_id, absence.person_id, absence.absence_day)
        ].add(absence.absence_type)

    if not data_dates:
        return []

    first_week = week_start(min(data_dates))
    last_week = week_start(max(data_dates))
    week_starts = [
        current for current in daterange(first_week, last_week) if current.weekday() == 0
    ]

    rows: list[dict[str, Any]] = []
    for (store_id, person_id), applicable in sorted(
        applicable_by_employee.items(), key=lambda item: (str(item[0][0]), str(item[0][1]))
    ):
        employee_key = (store_id, person_id)
        employee_work_values = worked_hours_by_employee.get(employee_key, [])
        average_daily_hours = (
            round(sum(employee_work_values) / len(employee_work_values), 4)
            if employee_work_values else None
        )
        presence_dates = employee_presence_dates.get(employee_key, set())
        absence_dates_all = {
            day for (sid, pid, day), types in absence_types_by_employee_day.items()
            if sid == store_id and pid == person_id and types
        }
        absent_entire_period = bool(
            presence_dates
            and presence_dates.issubset(absence_dates_all)
            and not worked_days_by_employee.get(employee_key)
        )

        for monday in week_starts:
            sunday = monday + timedelta(days=6)
            week_days = {monday + timedelta(days=i) for i in range(7)}
            covered_days = len(week_days & data_dates)
            complete_week = covered_days == 7
            planned = round(hours_by_employee_week.get((store_id, person_id, monday), 0.0), 4)

            try:
                contracted = float(applicable)
            except (TypeError, ValueError):
                contracted = None

            difference = round(planned - contracted, 4) if contracted is not None else None
            missing_hours = (
                round(max(contracted - planned, 0.0), 4) if contracted is not None else None
            )
            excess_hours = (
                round(max(planned - contracted, 0.0), 4) if contracted is not None else None
            )

            absence_days_without_shift = sorted(
                day for day in week_days
                if (store_id, person_id, day) in absence_types_by_employee_day
                and day not in worked_days_by_employee.get(employee_key, set())
                and day in data_dates
            )
            absence_types = sorted({
                absence_type
                for day in absence_days_without_shift
                for absence_type in absence_types_by_employee_day[(store_id, person_id, day)]
            })
            potential_absence_hours = (
                round(len(absence_days_without_shift) * average_daily_hours, 4)
                if average_daily_hours is not None else None
            )

            if not complete_week:
                status = "NO EVALUABLE"
            elif contracted is None:
                status = "SIN HORAS CONTRATO"
            elif abs(difference) <= WEEKLY_HOURS_TOLERANCE:
                status = "COINCIDE"
            elif difference < 0:
                status = "FALTAN HORAS"
            else:
                status = "EXCESO HORAS"

            has_missing_and_absence = bool(
                complete_week
                and missing_hours is not None
                and missing_hours > WEEKLY_HOURS_TOLERANCE
                and absence_days_without_shift
            )
            if absent_entire_period:
                absence_explanation = "AUSENTE TODO EL PERIODO"
            elif not has_missing_and_absence:
                absence_explanation = "NO"
            elif potential_absence_hours is None:
                absence_explanation = "AUSENCIA SIN MEDIA CALCULABLE"
            elif potential_absence_hours + WEEKLY_HOURS_TOLERANCE >= missing_hours:
                absence_explanation = "PODRIA EXPLICAR TODAS LAS HORAS FALTANTES"
            else:
                absence_explanation = "PODRIA EXPLICAR PARTE DE LAS HORAS FALTANTES"

            rows.append({
                "id_tienda": store_id,
                "personId": person_id,
                "ano_iso": monday.isocalendar().year,
                "semana_iso": monday.isocalendar().week,
                "inicio_semana": monday,
                "fin_semana": sunday,
                "dias_cubiertos_fichero": covered_days,
                "semana_completa_en_fichero": "SI" if complete_week else "NO",
                "applicableWorkingHours": contracted,
                "horas_planificadas": planned,
                "diferencia_planificadas_menos_contrato": difference,
                "horas_no_planificadas_hasta_contrato": missing_hours,
                "horas_planificadas_en_exceso": excess_hours,
                "estado_planificacion": status,
                "cumple_horas_contrato": (
                    "SI" if status == "COINCIDE" else
                    "NO" if status in {"FALTAN HORAS", "EXCESO HORAS"} else
                    "NO EVALUABLE"
                ),
                "media_horas_dia_planificado": average_daily_hours,
                "dias_ausencia_sin_turno": len(absence_days_without_shift),
                "fechas_ausencia_sin_turno": ", ".join(
                    day.strftime("%Y-%m-%d") for day in absence_days_without_shift
                ),
                "tipos_ausencia": ", ".join(absence_types),
                "horas_potenciales_asociadas_ausencia": potential_absence_hours,
                "faltan_horas_y_hay_ausencia": "SI" if has_missing_and_absence else "NO",
                "posible_explicacion_por_ausencia": absence_explanation,
                "ausente_todo_el_periodo": "SI" if absent_entire_period else "NO",
            })

    return rows

def analyze_shifts(
    shifts: list[ShiftRow], employee_months: dict[tuple[Any, Any, str], Any]
) -> tuple[list[dict[str, Any]], list[Incident]]:
    by_employee: dict[tuple[Any, Any], list[ShiftRow]] = defaultdict(list)
    for shift in shifts:
        by_employee[(shift.store_id, shift.person_id)].append(shift)

    incidents: list[Incident] = []

    # Incidencias de duracion y descanso entre jornadas.
    for (store_id, person_id), person_shifts in by_employee.items():
        person_shifts.sort(key=lambda row: (row.shift_start, row.shift_end))

        for shift in person_shifts:
            month = month_key(shift.work_day)
            if shift.worked_hours > MAX_SHIFT_HOURS:
                incidents.append(
                    Incident(
                        store_id, person_id, month, "TURNO_SUPERIOR_7_5H",
                        shift.work_day, shift.work_day, shift.worked_hours,
                        "<= 7,5 horas",
                        f"{shift.work_day:%d/%m/%Y}: {shift.worked_hours:.2f} h "
                        f"({shift.shift_start:%H:%M}-{shift.shift_end:%H:%M})",
                    )
                )
            if shift.worked_hours < MIN_SHIFT_HOURS:
                incidents.append(
                    Incident(
                        store_id, person_id, month, "TURNO_INFERIOR_4H",
                        shift.work_day, shift.work_day, shift.worked_hours,
                        ">= 4 horas",
                        f"{shift.work_day:%d/%m/%Y}: {shift.worked_hours:.2f} h "
                        f"({shift.shift_start:%H:%M}-{shift.shift_end:%H:%M})",
                    )
                )

        for previous, current in zip(person_shifts, person_shifts[1:]):
            rest_hours = (current.shift_start - previous.shift_end).total_seconds() / 3600
            if rest_hours < MIN_REST_HOURS:
                incidents.append(
                    Incident(
                        store_id, person_id, month_key(current.work_day),
                        "DESCANSO_INFERIOR_11H",
                        previous.work_day, current.work_day, round(rest_hours, 4),
                        ">= 11 horas",
                        f"{previous.work_day:%d/%m/%Y} {previous.shift_end:%H:%M} -> "
                        f"{current.work_day:%d/%m/%Y} {current.shift_start:%H:%M}: "
                        f"{rest_hours:.2f} h",
                    )
                )

        # Las rachas se calculan con todo el horizonte para detectar cruces de mes.
        for streak in find_consecutive_streaks([row.work_day for row in person_shifts]):
            if len(streak) <= MAX_CONSECUTIVE_DAYS:
                continue
            touched_months = sorted({month_key(day) for day in streak})
            for month in touched_months:
                incidents.append(
                    Incident(
                        store_id, person_id, month, "MAS_DE_5_DIAS_CONSECUTIVOS",
                        streak[0], streak[-1], float(len(streak)), "<= 5 dias",
                        f"{streak[0]:%d/%m/%Y}-{streak[-1]:%d/%m/%Y}: "
                        f"{len(streak)} dias consecutivos",
                    )
                )

    incident_groups: dict[tuple[Any, Any, str, str], list[Incident]] = defaultdict(list)
    for incident in incidents:
        incident_groups[
            (incident.store_id, incident.person_id, incident.month, incident.incident_type)
        ].append(incident)

    summaries: list[dict[str, Any]] = []
    for (store_id, person_id, month), applicable_hours in sorted(
        employee_months.items(), key=lambda item: (str(item[0][0]), str(item[0][1]), item[0][2])
    ):
        year, month_number = map(int, month.split("-"))
        month_shifts = [
            row for row in by_employee.get((store_id, person_id), [])
            if month_key(row.work_day) == month
        ]
        worked_days = {row.work_day for row in month_shifts}

        complete_weekends, free_saturdays, free_sundays = weekend_counts(
            year, month_number, worked_days
        )

        # Maximo de racha que toca el mes, aunque haya comenzado en el anterior.
        touching_streaks = [
            streak
            for streak in find_consecutive_streaks(
                [row.work_day for row in by_employee.get((store_id, person_id), [])]
            )
            if any(month_key(day) == month for day in streak)
        ]
        max_streak = max((len(streak) for streak in touching_streaks), default=0)

        count_consecutive = len(
            incident_groups.get((store_id, person_id, month, "MAS_DE_5_DIAS_CONSECUTIVOS"), [])
        )
        count_long = len(
            incident_groups.get((store_id, person_id, month, "TURNO_SUPERIOR_7_5H"), [])
        )
        count_short = len(
            incident_groups.get((store_id, person_id, month, "TURNO_INFERIOR_4H"), [])
        )
        count_rest = len(
            incident_groups.get((store_id, person_id, month, "DESCANSO_INFERIOR_11H"), [])
        )

        summaries.append(
            {
                "id_tienda": store_id,
                "personId": person_id,
                "applicableWorkingHours": applicable_hours,
                "mes": month,
                "dias_trabajados": len(worked_days),
                "max_dias_consecutivos": max_streak,
                "incidencias_dias_consecutivos": count_consecutive,
                "cumple_max_5_dias": "SI" if count_consecutive == 0 else "NO",
                "turnos_superiores_7_5h": count_long,
                "cumple_duracion_maxima": "SI" if count_long == 0 else "NO",
                "turnos_inferiores_4h": count_short,
                "cumple_duracion_minima": "SI" if count_short == 0 else "NO",
                "descansos_inferiores_11h": count_rest,
                "cumple_descanso_entre_jornadas": "SI" if count_rest == 0 else "NO",
                "cumple_todas_las_reglas": (
                    "SI" if count_consecutive + count_long + count_short + count_rest == 0 else "NO"
                ),
                "fines_semana_completos_libres": complete_weekends,
                "sabados_libres": free_saturdays,
                "domingos_libres": free_sundays,
            }
        )

    incidents.sort(
        key=lambda item: (
            str(item.store_id), str(item.person_id), item.month,
            item.start_date, item.incident_type
        )
    )
    return summaries, incidents


def style_header(worksheet, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[row].height = 32


def add_table(worksheet, name: str, last_column: str, row_count: int) -> None:
    if row_count <= 0:
        return
    table = Table(displayName=name, ref=f"A1:{last_column}{row_count + 1}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def write_flattened_sheet(workbook: Workbook, shifts: list[ShiftRow]) -> None:
    ws = workbook.active
    ws.title = "Planned Drafts"
    headers = [
        "id_tienda", "personId", "applicableWorkingHours", "day",
        "hora_inicio", "hora_fin", "horas_totales", "duracion_descanso",
    ]
    ws.append(headers)
    for row in shifts:
        ws.append([
            row.store_id,
            row.person_id,
            row.applicable_working_hours,
            row.work_day,
            row.shift_start.time(),
            row.shift_end.time(),
            row.worked_hours,
            row.break_hours,
        ])

    style_header(ws)
    ws.freeze_panes = "A2"
    for cell in ws["D"][1:]:
        cell.number_format = "yyyy-mm-dd"
    for column in ("E", "F"):
        for cell in ws[column][1:]:
            cell.number_format = "hh:mm"
    for column in ("C", "G", "H"):
        for cell in ws[column][1:]:
            cell.number_format = "0.00"

    widths = {"A": 13, "B": 13, "C": 24, "D": 13, "E": 13, "F": 13, "G": 15, "H": 20}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    add_table(ws, "PlannedDraftsTable", "H", len(shifts))


def write_summary_sheet(workbook: Workbook, summaries: list[dict[str, Any]]) -> None:
    ws = workbook.create_sheet("Validacion mensual")
    headers = [
        "id_tienda", "personId", "applicableWorkingHours", "mes",
        "dias_trabajados", "max_dias_consecutivos", "incidencias_dias_consecutivos",
        "cumple_max_5_dias", "turnos_superiores_7_5h", "cumple_duracion_maxima",
        "turnos_inferiores_4h", "cumple_duracion_minima",
        "descansos_inferiores_11h", "cumple_descanso_entre_jornadas",
        "cumple_todas_las_reglas", "fines_semana_completos_libres",
        "sabados_libres", "domingos_libres",
    ]
    ws.append(headers)
    for row in summaries:
        ws.append([row[header] for header in headers])

    style_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "A": 12, "B": 13, "C": 23, "D": 11, "E": 15, "F": 22,
        "G": 25, "H": 20, "I": 23, "J": 23, "K": 21, "L": 23,
        "M": 24, "N": 29, "O": 22, "P": 31, "Q": 16, "R": 17,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    red_fill = PatternFill("solid", fgColor="F4CCCC")
    for col in ("H", "J", "L", "N", "O"):
        ws.conditional_formatting.add(
            f"{col}2:{col}{max(len(summaries) + 1, 2)}",
            FormulaRule(formula=[f'{col}2="NO"'], fill=red_fill),
        )

    add_table(ws, "ValidacionMensualTable", "R", len(summaries))


def write_detail_sheet(workbook: Workbook, incidents: list[Incident]) -> None:
    ws = workbook.create_sheet("Detalle incidencias")
    headers = [
        "id_tienda", "personId", "mes", "tipo_incidencia", "fecha_inicio",
        "fecha_fin", "valor_observado", "limite", "detalle",
    ]
    ws.append(headers)
    for item in incidents:
        ws.append([
            item.store_id,
            item.person_id,
            item.month,
            item.incident_type,
            item.start_date,
            item.end_date,
            item.observed_value,
            item.limit_text,
            item.detail,
        ])

    style_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in ("E", "F"):
        for cell in ws[column][1:]:
            cell.number_format = "yyyy-mm-dd"
    for cell in ws["G"][1:]:
        cell.number_format = "0.00"

    widths = {
        "A": 12, "B": 13, "C": 11, "D": 31, "E": 14,
        "F": 14, "G": 17, "H": 18, "I": 58,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2, min_col=9, max_col=9):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")

    add_table(ws, "DetalleIncidenciasTable", "I", len(incidents))



def write_weekly_hours_sheet(workbook: Workbook, weekly_rows: list[dict[str, Any]]) -> None:
    ws = workbook.create_sheet("Control horas semanal")
    headers = [
        "id_tienda", "personId", "ano_iso", "semana_iso",
        "inicio_semana", "fin_semana", "dias_cubiertos_fichero",
        "semana_completa_en_fichero", "applicableWorkingHours",
        "horas_planificadas", "diferencia_planificadas_menos_contrato",
        "horas_no_planificadas_hasta_contrato", "horas_planificadas_en_exceso",
        "estado_planificacion", "cumple_horas_contrato",
        "media_horas_dia_planificado", "dias_ausencia_sin_turno",
        "fechas_ausencia_sin_turno", "tipos_ausencia",
        "horas_potenciales_asociadas_ausencia",
        "faltan_horas_y_hay_ausencia", "posible_explicacion_por_ausencia",
        "ausente_todo_el_periodo",
    ]
    ws.append(headers)
    for row in weekly_rows:
        ws.append([row[header] for header in headers])

    style_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column in ("E", "F"):
        for cell in ws[column][1:]:
            cell.number_format = "yyyy-mm-dd"
    for column in ("I", "J", "K", "L", "M", "P", "T"):
        for cell in ws[column][1:]:
            cell.number_format = "0.00"

    widths = {
        "A": 12, "B": 13, "C": 10, "D": 12, "E": 15, "F": 15,
        "G": 23, "H": 28, "I": 24, "J": 20, "K": 39, "L": 35,
        "M": 30, "N": 22, "O": 25, "P": 27, "Q": 25, "R": 40,
        "S": 26, "T": 38, "U": 28, "V": 44, "W": 25,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    red_fill = PatternFill("solid", fgColor="F4CCCC")
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    green_fill = PatternFill("solid", fgColor="D9EAD3")
    blue_fill = PatternFill("solid", fgColor="D9EAF7")
    last_row = max(len(weekly_rows) + 1, 2)
    ws.conditional_formatting.add(
        f"N2:N{last_row}", FormulaRule(formula=['OR(N2="FALTAN HORAS",N2="EXCESO HORAS")'], fill=red_fill)
    )
    ws.conditional_formatting.add(
        f"N2:N{last_row}", FormulaRule(formula=['OR(N2="NO EVALUABLE",N2="SIN HORAS CONTRATO")'], fill=yellow_fill)
    )
    ws.conditional_formatting.add(
        f"N2:N{last_row}", FormulaRule(formula=['N2="COINCIDE"'], fill=green_fill)
    )
    ws.conditional_formatting.add(
        f"V2:V{last_row}", FormulaRule(formula=['V2<>"NO"'], fill=blue_fill)
    )
    ws.conditional_formatting.add(
        f"W2:W{last_row}", FormulaRule(formula=['W2="SI"'], fill=yellow_fill)
    )

    for row in ws.iter_rows(min_row=2, min_col=18, max_col=22):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    add_table(ws, "ControlHorasSemanalTable", "W", len(weekly_rows))

def write_excel(
    shifts: list[ShiftRow],
    summaries: list[dict[str, Any]],
    incidents: list[Incident],
    weekly_rows: list[dict[str, Any]],
    output_path: Path,
) -> None:
    workbook = Workbook()
    write_flattened_sheet(workbook, shifts)
    write_summary_sheet(workbook, summaries)
    write_detail_sheet(workbook, incidents)
    write_weekly_hours_sheet(workbook, weekly_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        workbook.save(output_path)
    except PermissionError as exc:
        raise PermissionError(
            f"No se pudo guardar '{output_path}'. Comprueba que el Excel no "
            "este abierto y que tienes permisos de escritura."
        ) from exc


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Aplana los plannedDraft del JSON, valida las restricciones de "
            "turnos y genera un Excel con resumen mensual, detalle y control semanal de horas."
        )
    )
    parser.add_argument(
        "input_txt", nargs="?", default=DEFAULT_INPUT,
        help=f"Fichero JSON/TXT de entrada. Por defecto: {DEFAULT_INPUT}",
    )
    parser.add_argument(
        "output_xlsx", nargs="?", default=DEFAULT_OUTPUT,
        help=f"Excel de salida. Por defecto: {DEFAULT_OUTPUT}",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()
    input_path = Path(args.input_txt).expanduser().resolve()
    output_path = Path(args.output_xlsx).expanduser().resolve()

    try:
        data = load_json(input_path)
        shifts, employee_months, absences, employee_presence_dates = extract_data(data)
        summaries, incidents = analyze_shifts(shifts, employee_months)
        data_dates = {
            date.fromisoformat(str(item.get("operatingDate"))[:10])
            for item in (data.get("storeDayTimes") or [])
            if isinstance(item, dict) and item.get("operatingDate")
        }
        weekly_rows = analyze_weekly_hours(
            shifts, employee_months, data_dates, absences, employee_presence_dates
        )
        write_excel(shifts, summaries, incidents, weekly_rows, output_path)
    except (FileNotFoundError, ValueError, PermissionError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    except Exception as exc:
        print(f"ERROR INESPERADO: {exc}", file=sys.stderr)
        return 1

    print(f"Excel generado correctamente: {output_path}")
    print(f"Turnos aplanados: {len(shifts)}")
    print(f"Registros mensuales: {len(summaries)}")
    print(f"Incidencias detalladas: {len(incidents)}")
    print(f"Controles semanales: {len(weekly_rows)}")
    print(f"Dias de ausencia detectados: {len(absences)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
